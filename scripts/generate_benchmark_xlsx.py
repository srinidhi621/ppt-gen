"""
Generate the V3 benchmark test prompts Excel file.

Usage:
    python scripts/generate_benchmark_xlsx.py

Output:
    assets/benchmarks/v3_test_prompts.xlsx

Rerun this script after editing the TEST_PROMPTS or AXIS_DEFINITIONS
data structures below. The Excel is fully regenerated each time.
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

OUTPUT_PATH = Path(__file__).resolve().parent.parent / "assets" / "benchmarks" / "v3_test_prompts.xlsx"

# ── Evaluation axes ──────────────────────────────────────────────

AXIS_DEFINITIONS = [
    {
        "axis": "Content Fidelity",
        "1": "Key content missing or fabricated. User would discard.",
        "2": "Most content present but important items wrong or missing.",
        "3": "All user-specified content present. Minor rewordings acceptable.",
        "4": "Content faithfully represented with appropriate emphasis.",
        "5": "Content perfectly captured; emphasis and framing add value beyond the prompt.",
    },
    {
        "axis": "Archetype Selection",
        "1": "Wrong layout family entirely (e.g. KPI grid for a timeline).",
        "2": "Plausible family but poor fit for the content structure.",
        "3": "Correct family. Layout works but a better option existed.",
        "4": "Right family, well-suited to the content and audience.",
        "5": "Optimal family choice; hard to argue for an alternative.",
    },
    {
        "axis": "Visual Hierarchy",
        "1": "No clear focal point. Reader does not know where to look.",
        "2": "Focal point exists but competes with secondary elements.",
        "3": "Clear primary element. Supporting content is readable.",
        "4": "Strong hierarchy. Eye flows naturally from headline to evidence to detail.",
        "5": "Magazine-quality flow. Hierarchy guides comprehension effortlessly.",
    },
    {
        "axis": "Density & Readability",
        "1": "Text overflows, overlaps, or is unreadably small.",
        "2": "Readable but too dense or too sparse. Whitespace badly distributed.",
        "3": "Content fits. Whitespace is reasonable. No overflow.",
        "4": "Well-balanced density. Breathing room without wasted space.",
        "5": "Density feels intentional. Every element earns its space.",
    },
    {
        "axis": "Brand Consistency",
        "1": "Wrong colors, wrong fonts, or both. Does not look branded.",
        "2": "Mostly branded but 1-2 off-palette elements or wrong font.",
        "3": "All elements use brand tokens. No palette drift.",
        "4": "Brand-consistent and the brand treatment enhances the message.",
        "5": "Indistinguishable from a designer-authored branded slide.",
    },
    {
        "axis": "Editability",
        "1": "Shapes are grouped/locked, text not selectable, or layout breaks on edit.",
        "2": "Editable but structure is fragile — moving one element breaks alignment.",
        "3": "All text and shapes editable. Layout holds with minor content changes.",
        "4": "Easy to edit. Guides/grid make it clear how to modify.",
        "5": "A non-designer could confidently edit content and maintain quality.",
    },
    {
        "axis": "Mechanical Defects",
        "1": "Multiple blocking defects (overflow, off-canvas, broken images).",
        "2": "One blocking defect or multiple warnings.",
        "3": "No blocking defects. Minor warnings (slight overlap, font fallback).",
        "4": "Clean scan. No blocking defects, no warnings.",
        "5": "Perfect mechanical execution. Scanner returns zero findings.",
    },
    {
        "axis": "Cross-Slide Consistency (multi-slide only)",
        "1": "Slides look like they came from different decks. No visual thread.",
        "2": "Some consistency but jarring shifts in style or spacing between slides.",
        "3": "Consistent token usage. Minor drift in spacing or hierarchy treatment.",
        "4": "Cohesive deck. Shared rhythm, accent strategy, and type hierarchy throughout.",
        "5": "Feels like one designer authored the entire deck in one sitting.",
    },
    {
        "axis": "Narrative Flow (multi-slide only)",
        "1": "Slide order makes no sense. No argument spine.",
        "2": "Logical order but transitions are abrupt or repetitive.",
        "3": "Clear progression. Each slide advances the argument.",
        "4": "Strong narrative arc. Audience builds understanding slide by slide.",
        "5": "Compelling storytelling. The deck persuades through structure alone.",
    },
]

# ── Test prompts ─────────────────────────────────────────────────
# Each tuple: (test_id, category, target_archetypes, user_instruction,
#               what_system_must_figure_out, expected_slides, notes)

TEST_PROMPTS = [
    # ── Single-slide: core archetypes ────────────────────────────
    (
        "TP-01", "Executive Opener", "hero_title",
        "I need a title slide for a modernization pitch to the CTO of a large bank. "
        "Something bold. The program is called 'Legacy Navigator'.",
        "Audience (CTO/executive), tone (bold but not flashy), "
        "that 'Legacy Navigator' is a program name not a generic phrase, "
        "appropriate subtitle generation",
        "1",
        "Should produce a single hero slide with headline + subtitle. "
        "Test: does it avoid generic AI-smell language?",
    ),
    (
        "TP-02", "Strategic Thesis", "hero_statement_with_support_columns",
        "We need to show that legacy fragmentation is hurting delivery speed. "
        "Three reasons: duplicate systems, manual workarounds, and talent drain. "
        "Make it punchy.",
        "That 'punchy' means concise not loud, that the three reasons are the "
        "support columns, appropriate density for an executive audience",
        "1",
        "Core test for the thesis+supports archetype. "
        "Checks density budget, hierarchy, one-claim-per-slide.",
    ),
    (
        "TP-03", "Capability Overview", "three_cards",
        "Show our three AI service lines: AI Consulting & Strategy, "
        "AI-Led Process Reimagination, and AI-Assisted Software Engineering. "
        "Keep descriptions short, we'll present verbally.",
        "That 'keep descriptions short' implies low density, "
        "that each service line is a card, not bullets, "
        "that verbal delivery means less text on slide",
        "1",
        "Three-card layout test. Checks equal weighting, "
        "accent usage, brevity.",
    ),
    (
        "TP-04", "Current vs Future", "comparison_split",
        "Compare where the client is today versus where we'll take them. "
        "Today: siloed data, manual reporting, 14-day change lead time. "
        "Future: unified platform, real-time dashboards, 3-day cycles.",
        "That this is a two-panel comparison, "
        "that the numbers are key evidence and should be prominent, "
        "visual contrast between 'bad now' and 'good future'",
        "1",
        "Split layout test. Checks visual contrast, "
        "number prominence, balanced panels.",
    ),
    (
        "TP-05", "Delivery Metrics", "kpi_grid",
        "Put together a metrics slide. We reduced incidents by 40%, "
        "cut deployment time from 2 weeks to 2 days, "
        "onboarded 150 engineers, and saved $2.3M annually. "
        "Maybe add one more if it looks empty.",
        "That 'maybe add one more' is a density judgment call (4 is fine), "
        "that the numbers are the focal point not the labels, "
        "appropriate formatting for currency vs percentage vs count",
        "1",
        "KPI grid test. Checks number hierarchy, "
        "label sizing, even spacing.",
    ),
    (
        "TP-06", "Methodology / Approach", "process_flow",
        "Walk them through our modernization approach. "
        "It goes: Discover, Assess, Stabilize, Migrate, Optimize. "
        "Each step should have a one-liner underneath.",
        "That 5 steps is within process_flow capacity, "
        "that 'one-liner' means max ~10 words per caption, "
        "horizontal flow with connectors",
        "1",
        "Process flow test. Checks connector styling, "
        "step spacing, caption brevity.",
    ),
    (
        "TP-07", "Program Roadmap", "timeline_roadmap",
        "Show the 18-month delivery roadmap. "
        "Phase 1 is foundation (3 months), Phase 2 is migration (6 months), "
        "Phase 3 is optimization (6 months), Phase 4 is hypercare (3 months). "
        "Mark where the first production release happens.",
        "That this is a timeline not a process flow, "
        "that durations should be visually proportional or labeled, "
        "that the production release is a milestone marker, "
        "that 4 phases with a milestone is 5 visual elements",
        "1",
        "Timeline test. Checks phase proportionality, "
        "milestone callout, duration labels.",
    ),
    (
        "TP-08", "Assessment Matrix", "matrix_grid",
        "We did an AI readiness assessment across four areas: "
        "Data Infrastructure, Talent & Skills, Governance, and Technology Stack. "
        "For each, show the current maturity, key gap, and our recommendation. "
        "It's going to a steering committee so keep it dense but scannable.",
        "That this is a 4-row x 3-column matrix, "
        "that 'dense but scannable' means structured text not paragraphs, "
        "appropriate for steering committee (formal, evidence-heavy)",
        "1",
        "Matrix grid test. Checks cell alignment, "
        "header clarity, manageable density.",
    ),
    (
        "TP-09", "Solution Concept", "content_with_visual",
        "Show what the modernized platform looks like. "
        "Left side should explain the key components in 3-4 bullets. "
        "Right side needs some kind of visual — a simplified architecture "
        "or a conceptual diagram. Nothing too detailed.",
        "That the visual side needs a generated or placeholder diagram, "
        "that 'nothing too detailed' constrains visual complexity, "
        "that 3-4 bullets is the density cap for the text side",
        "1",
        "Split content+visual test. Checks panel balance, "
        "visual placeholder handling, bullet discipline.",
    ),
    (
        "TP-10", "Next Steps / CTA", "closing_cta",
        "End the deck with next steps. We need a workshop scheduled within 2 weeks, "
        "a technical deep-dive with their platform team, "
        "and a proposal by end of month. Make it action-oriented.",
        "That 'action-oriented' means imperative verbs not passive descriptions, "
        "that 3 items is within closing_cta capacity, "
        "that dates/deadlines should be prominent",
        "1",
        "Closing CTA test. Checks action verb usage, "
        "deadline prominence, clean closure feel.",
    ),
    # ── Single-slide: untested archetypes ────────────────────────
    (
        "TP-11", "Client Testimonial", "quote_callout",
        "We have a great quote from the VP of Engineering at a pharma client: "
        "'Ascendion's team didn't just modernize our stack — they changed how our "
        "engineers think about delivery.' Put it on a slide. "
        "Attribution: Rajesh Menon, VP Engineering, NovaPharma.",
        "That this is a single-quote slide not a bullet list, "
        "that the attribution needs name + title + company, "
        "that the quote should be the dominant visual element",
        "1",
        "Quote callout archetype. Checks quote prominence, "
        "attribution formatting, elegant whitespace.",
    ),
    (
        "TP-12", "Section Divider", "section_break",
        "I need a divider slide before we get into the technical details. "
        "Something like 'Our Approach' or 'How We'll Get There'. "
        "Don't overthink it.",
        "That 'don't overthink it' means minimal content, "
        "that the user wants a mood/transition not information, "
        "strong visual treatment with minimal text",
        "1",
        "Section break archetype. Checks visual weight, "
        "minimal text, transition feel.",
    ),
    (
        "TP-13", "Impact Stats with Icons", "stat_list_with_icons",
        "List out the key results from the engagement: 40% fewer incidents, "
        "3x faster deployments, 98.5% platform uptime, "
        "$2.3M annual savings, 150 engineers onboarded. "
        "Use icons if you have them.",
        "That 5 stats maps to stat_list_with_icons not kpi_grid, "
        "that 'use icons if you have them' is conditional, "
        "that this is a vertical list not a grid layout",
        "1",
        "Stat list archetype. Checks icon-to-stat alignment, "
        "vertical rhythm, number formatting. Compare against TP-05 "
        "which has similar content but different archetype.",
    ),
    # ── Single-slide: edge cases ─────────────────────────────────
    (
        "TP-14", "Sparse Content", "hero_statement_with_support_columns",
        "Just put up 'We simplify legacy.' with two supporting points: "
        "clarity and speed. That's it.",
        "That the content is intentionally minimal, "
        "that 2 supports (not 3) requires layout adjustment, "
        "that the slide should NOT look empty — whitespace is intentional",
        "1",
        "Edge case: under-filled archetype. Tests whether the system "
        "handles sparse content gracefully instead of stretching or padding.",
    ),
    (
        "TP-15", "Dense Content at Capacity Limit", "matrix_grid",
        "Show the full assessment results. Rows: Data Infrastructure, "
        "Application Portfolio, Security & Compliance, Talent & Skills. "
        "Columns: Current State, Gap Analysis, Recommended Actions, Timeline, Owner. "
        "Each cell needs 2-3 sentences.",
        "That 4 rows x 5 columns with 2-3 sentences per cell EXCEEDS "
        "matrix_grid capacity (max 4x3). The feasibility gate should trigger. "
        "System must reduce columns, split, or push back.",
        "1",
        "Edge case: content exceeds archetype capacity. "
        "Tests feasibility gate. Should NOT attempt to cram — should re-plan.",
    ),
    (
        "TP-16", "Ambiguous Archetype", "process_flow OR timeline_roadmap",
        "Show how the program unfolds over the next year. "
        "Discovery in Q1, Build in Q2-Q3, Launch in Q4. "
        "Each phase has key deliverables.",
        "That this could be a process_flow (steps) or timeline_roadmap "
        "(temporal). Either is defensible. System should pick one "
        "and commit, not produce a hybrid.",
        "1",
        "Edge case: ambiguous archetype. Tests whether the planner "
        "makes a clean choice. Either archetype is acceptable; "
        "a confused hybrid is not.",
    ),
    # ── Single-slide: audience variations ────────────────────────
    (
        "TP-17", "Board-Level Summary", "kpi_grid",
        "The board wants a one-page summary of the AI program outcomes. "
        "Revenue impact: +$4.2M. Cost reduction: $1.8M. "
        "Client NPS: 72 (up from 58). Team utilization: 91%. "
        "Keep it at the level a board member reads in 5 seconds.",
        "That 'board member reads in 5 seconds' means VERY sparse, "
        "numbers must be enormous, labels minimal, "
        "no supporting detail whatsoever",
        "1",
        "Audience test: board-level. Same archetype as TP-05 but "
        "radically different density. Checks audience-adaptive styling.",
    ),
    (
        "TP-18", "Technical Team Detail", "matrix_grid",
        "For the engineering leads meeting, show the migration status "
        "per service: Auth Service (migrated, on k8s, 99.9% uptime), "
        "Payment Gateway (in progress, legacy DB dependency, blocked on schema), "
        "Notification Hub (planned, depends on Auth), "
        "Analytics Pipeline (migrated, Spark on EMR, monitoring gap). "
        "They want the detail.",
        "That 'they want the detail' means higher density is appropriate, "
        "that status indicators (migrated/in-progress/planned) need visual coding, "
        "technical audience tolerates and expects data-rich slides",
        "1",
        "Audience test: technical team. Dense matrix is appropriate here. "
        "Checks status-indicator styling, technical terminology handling.",
    ),
    # ── Single-slide: content type variations ────────────────────
    (
        "TP-19", "Case Study Narrative", "content_with_visual",
        "Tell the NovaPharma story. They had 200+ legacy services, "
        "18-month release cycles, and were losing engineers. "
        "We consolidated to 40 services on a modern platform, "
        "cut releases to biweekly, and attrition dropped 30%. "
        "Show the before/after somehow.",
        "That this is a narrative not a data slide, "
        "that 'show the before/after somehow' could mean a comparison "
        "visual or a transformation arrow, "
        "that the numbers are evidence supporting a story not the story itself",
        "1",
        "Content type: narrative + evidence. Tests whether the builder "
        "balances storytelling with data. Could also test comparison_split; "
        "content_with_visual is the target.",
    ),
    (
        "TP-20", "Sales-Oriented Persuasion", "three_cards",
        "We're pitching to a new prospect. Show our three differentiators: "
        "we've done this 50+ times, we have 1100+ AI-certified engineers, "
        "and our average engagement NPS is 74. "
        "Make it compelling, not just informative.",
        "That 'compelling not informative' means the framing matters more "
        "than the data, that numbers should support a claim not be the claim, "
        "that sales tone is confident but not pushy",
        "1",
        "Content type: persuasive. Same archetype as TP-03 but "
        "sales-oriented. Checks tone calibration.",
    ),
    # ── Multi-slide: deck-level tests ────────────────────────────
    (
        "TP-21", "Executive Pitch Deck", "hero_title, hero_statement, "
        "three_cards, process_flow, kpi_grid, closing_cta",
        "Build me a short pitch deck for legacy modernization. "
        "We're presenting to a bank's CTO and Head of Engineering. "
        "Cover: who we are, why legacy is a problem, our approach, "
        "what we've delivered before, and next steps. "
        "Keep it to 5-6 slides. Executive audience.",
        "Slide count (5-6), archetype selection across the deck, "
        "narrative arc (problem → approach → proof → action), "
        "cross-slide visual consistency, audience-appropriate density, "
        "variety (no two consecutive slides with the same layout)",
        "5-6",
        "Core multi-slide test. Exercises the full planner: archetype "
        "selection, narrative ordering, density calibration, variety. "
        "Score on Cross-Slide Consistency and Narrative Flow axes.",
    ),
    (
        "TP-22", "Client Proposal", "hero_title, comparison_split, "
        "matrix_grid, process_flow, timeline_roadmap, kpi_grid, "
        "closing_cta",
        "Put together a proposal deck for Globe Telecom. "
        "They want AI readiness assessment + implementation roadmap. "
        "Cover the problem, our assessment framework, what we'll find, "
        "how we'll fix it, timeline, proof points from similar work, "
        "and commercial next steps. Maybe 7-8 slides.",
        "That this is a proposal (persuasive + structured), "
        "that 7-8 slides requires careful archetype distribution, "
        "that 'assessment framework' likely maps to matrix_grid, "
        "that 'proof points' could be kpi_grid or stat_list, "
        "narrative must build from problem to solution to proof to ask",
        "7-8",
        "Dense multi-slide test. Exercises archetype variety across "
        "a longer deck. Tests whether the system avoids repetitive layouts "
        "and maintains coherence over 7-8 slides.",
    ),
    (
        "TP-23", "Quarterly Business Review", "hero_title, kpi_grid, "
        "comparison_split, process_flow, matrix_grid, closing_cta",
        "I need a QBR deck for the leadership team. Cover: program status "
        "(green/amber/red by workstream), key metrics for the quarter, "
        "what went well vs. what didn't, plan for next quarter, "
        "risks and mitigations, and decisions we need from leadership. "
        "6 slides max.",
        "That 'program status' with RAG ratings maps to matrix_grid, "
        "that 'what went well vs. what didn't' is a comparison_split, "
        "that 'decisions we need' is a closing_cta variant, "
        "QBR tone is factual and operational, not persuasive",
        "6",
        "Multi-slide test: operational/status genre. Different tone "
        "from TP-21 (pitch) and TP-22 (proposal). Tests whether the "
        "system adapts style_contract to operational context.",
    ),
    (
        "TP-24", "All-Hands Update", "hero_title, hero_statement, "
        "kpi_grid, quote_callout, timeline_roadmap, closing_cta",
        "Quick all-hands update for the AI practice. We hit our Q1 targets, "
        "team grew from 80 to 150, landed 3 new clients, NPS is at 74. "
        "Include that quote from the NovaPharma VP. "
        "Show what's coming in Q2. Keep it energetic. 5 slides.",
        "That all-hands tone is energetic and informal (not boardroom), "
        "that the quote from TP-11 should be reusable here, "
        "that 'what's coming in Q2' is a timeline_roadmap, "
        "5 slides means some content must be combined or cut",
        "5",
        "Multi-slide test: internal/informal genre. Tests tone shift "
        "from external-facing decks. Also tests quote_callout in "
        "a multi-slide context.",
    ),
    # ── Multi-slide: stress tests ────────────────────────────────
    (
        "TP-25", "Minimal Prompt / Maximum Ambiguity",
        "system must decide",
        "Make a deck about our AI consulting practice. "
        "For a potential client. 5 slides or so.",
        "EVERYTHING. No content provided — system must generate. "
        "Archetype selection, content invention, density choices, "
        "narrative structure, all from a 3-sentence prompt. "
        "This is the hardest test.",
        "~5",
        "Stress test: minimal input. The system has to make all decisions. "
        "Acceptable if the output is a reasonable generic pitch. "
        "Fails if it hallucinates specific numbers or makes up case studies.",
    ),
    (
        "TP-26", "Content Dump / No Structure Given",
        "system must decide",
        "Here's what I want to cover: we have 1100 AI engineers, "
        "we've done 50+ engagements, our NPS is 74, we offer consulting "
        "strategy and process reimagination and software engineering and "
        "talent solutions, our approach is discover then build then optimize, "
        "we partnered with AWS and Azure and Google Cloud, "
        "the timeline is usually 4-6 months for phase 1, "
        "and we have case studies from pharma banking and telecom. "
        "Turn this into a deck.",
        "That the user has given a content dump with no structure, "
        "the system must impose narrative order, "
        "group related content into slides, "
        "choose archetypes that fit each grouping, "
        "respect density limits (this content won't fit on 3 slides)",
        "6-8",
        "Stress test: unstructured input. Tests whether the planner "
        "can impose narrative structure on a flat content dump. "
        "The content is real but the organization is absent.",
    ),
]


# ── Excel generation ─────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def style_header_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BORDER


def style_body_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def add_section_row(ws, row, num_cols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BOLD_FONT
    cell.fill = SECTION_FILL
    cell.alignment = WRAP
    cell.border = THIN_BORDER
    for c in range(2, num_cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = SECTION_FILL


def build_prompts_sheet(wb):
    ws = wb.active
    ws.title = "Test Prompts"

    headers = [
        "Test ID", "Category", "Target Archetype(s)",
        "User Instruction (deliberately ambiguous)",
        "What the system must figure out",
        "Expected Slides", "Notes",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    sections = [
        ("Single-slide: core archetypes", "TP-01", "TP-10"),
        ("Single-slide: untested archetypes", "TP-11", "TP-13"),
        ("Single-slide: edge cases", "TP-14", "TP-16"),
        ("Single-slide: audience variations", "TP-17", "TP-18"),
        ("Single-slide: content type variations", "TP-19", "TP-20"),
        ("Multi-slide: deck-level tests", "TP-21", "TP-24"),
        ("Multi-slide: stress tests", "TP-25", "TP-26"),
    ]

    row = 2
    section_idx = 0
    for prompt in TEST_PROMPTS:
        tid = prompt[0]
        if section_idx < len(sections) and tid == sections[section_idx][1]:
            add_section_row(ws, row, len(headers), sections[section_idx][0])
            row += 1
            section_idx += 1

        for c, val in enumerate(prompt, 1):
            ws.cell(row=row, column=c, value=val)
        style_body_row(ws, row, len(headers))
        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 42

    for r in range(2, row):
        ws.row_dimensions[r].height = 70

    return ws


def build_rubric_sheet(wb):
    ws = wb.create_sheet("Evaluation Rubric")

    score_axes = [
        "Content\nFidelity", "Archetype\nSelection", "Visual\nHierarchy",
        "Density &\nReadability", "Brand\nConsistency", "Editability",
        "Mechanical\nDefects", "Cross-Slide\nConsistency", "Narrative\nFlow",
    ]
    headers = ["Test ID", "Slides"] + score_axes + ["Average", "Overall\nPass?", "Reviewer Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    sections = [
        ("Single-slide: core archetypes", "TP-01", "TP-10"),
        ("Single-slide: untested archetypes", "TP-11", "TP-13"),
        ("Single-slide: edge cases", "TP-14", "TP-16"),
        ("Single-slide: audience variations", "TP-17", "TP-18"),
        ("Single-slide: content type variations", "TP-19", "TP-20"),
        ("Multi-slide: deck-level tests", "TP-21", "TP-24"),
        ("Multi-slide: stress tests", "TP-25", "TP-26"),
    ]

    row = 2
    section_idx = 0
    num_score_cols = len(score_axes)
    first_score_col = 3
    last_score_col = first_score_col + num_score_cols - 1
    avg_col = last_score_col + 1
    pass_col = avg_col + 1
    notes_col = pass_col + 1

    for prompt in TEST_PROMPTS:
        tid = prompt[0]
        expected = prompt[5]
        is_multi = expected != "1"

        if section_idx < len(sections) and tid == sections[section_idx][1]:
            add_section_row(ws, row, len(headers), sections[section_idx][0])
            row += 1
            section_idx += 1

        ws.cell(row=row, column=1, value=tid).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=expected).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).border = THIN_BORDER

        for c in range(first_score_col, last_score_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            axis_idx = c - first_score_col
            if not is_multi and axis_idx >= 7:
                cell.value = "n/a"
                cell.font = Font(name="Calibri", size=9, italic=True, color="999999")

        # Average: only count numeric cells (skip n/a for single-slide)
        fc = get_column_letter(first_score_col)
        lc = get_column_letter(last_score_col)
        avg_cell = ws.cell(row=row, column=avg_col)
        avg_cell.value = f"=AVERAGE({fc}{row}:{lc}{row})"
        avg_cell.font = BOLD_FONT
        avg_cell.alignment = Alignment(horizontal="center", vertical="center")
        avg_cell.border = THIN_BORDER
        avg_cell.number_format = "0.0"

        # Pass: avg >= 3.5 AND no scored axis <= 2
        # For single-slide, only check first 7 axes
        if is_multi:
            min_range = f"{fc}{row}:{lc}{row}"
        else:
            lc7 = get_column_letter(first_score_col + 6)
            min_range = f"{fc}{row}:{lc7}{row}"

        pass_cell = ws.cell(row=row, column=pass_col)
        acol = get_column_letter(avg_col)
        pass_cell.value = (
            f'=IF(AND({acol}{row}>=3.5,MIN({min_range})>2),"PASS",'
            f'IF(ISBLANK({fc}{row}),"","FAIL"))'
        )
        pass_cell.font = BOLD_FONT
        pass_cell.alignment = Alignment(horizontal="center", vertical="center")
        pass_cell.border = THIN_BORDER

        notes_cell = ws.cell(row=row, column=notes_col)
        notes_cell.font = BODY_FONT
        notes_cell.alignment = WRAP
        notes_cell.border = THIN_BORDER

        row += 1

    # Summary rows
    row += 1
    ws.cell(row=row, column=1, value="Pass Criteria:").font = BOLD_FONT
    ws.cell(row=row, column=2, value=(
        "Per-prompt: average >= 3.5 across all scored axes AND no single axis <= 2. "
        "Mirrors SPEC-v3 quality gates."
    )).font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=notes_col)

    row += 1
    ws.cell(row=row, column=1, value="Benchmark Pass:").font = BOLD_FONT
    ws.cell(row=row, column=2, value=(
        "V3 ships as default when >= 70% of test prompts pass AND V3 output is "
        "rated higher than V1 on a majority of prompts."
    )).font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=notes_col)

    row += 1
    ws.cell(row=row, column=1, value="Multi-slide axes:").font = BOLD_FONT
    ws.cell(row=row, column=2, value=(
        "Cross-Slide Consistency and Narrative Flow are scored only for multi-slide "
        "tests (TP-21 through TP-26). Single-slide tests show 'n/a' for those columns."
    )).font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=notes_col)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 7
    for col_idx in range(first_score_col, last_score_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    ws.column_dimensions[get_column_letter(avg_col)].width = 9
    ws.column_dimensions[get_column_letter(pass_col)].width = 9
    ws.column_dimensions[get_column_letter(notes_col)].width = 45

    for r in range(2, row):
        if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < 30:
            ws.row_dimensions[r].height = 30

    return ws


def build_axis_definitions_sheet(wb):
    ws = wb.create_sheet("Axis Definitions")

    headers = ["Axis", "Score 1 (Fail)", "Score 2 (Poor)",
               "Score 3 (Acceptable)", "Score 4 (Good)", "Score 5 (Excellent)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for r, axis_def in enumerate(AXIS_DEFINITIONS, 2):
        ws.cell(row=r, column=1, value=axis_def["axis"])
        ws.cell(row=r, column=2, value=axis_def["1"])
        ws.cell(row=r, column=3, value=axis_def["2"])
        ws.cell(row=r, column=4, value=axis_def["3"])
        ws.cell(row=r, column=5, value=axis_def["4"])
        ws.cell(row=r, column=6, value=axis_def["5"])
        style_body_row(ws, r, len(headers))
        ws.row_dimensions[r].height = 65

    # Note about multi-slide axes
    note_row = len(AXIS_DEFINITIONS) + 3
    ws.cell(row=note_row, column=1, value="Note:").font = BOLD_FONT
    ws.cell(row=note_row, column=2, value=(
        "Cross-Slide Consistency and Narrative Flow are scored only for multi-slide "
        "tests. For single-slide tests, these columns show 'n/a' and are excluded from "
        "the average calculation."
    )).font = BODY_FONT
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=6)

    ws.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 35

    return ws


def get_column_letter(col_idx):
    from openpyxl.utils import get_column_letter as gcl
    return gcl(col_idx)


def main():
    wb = openpyxl.Workbook()
    build_prompts_sheet(wb)
    build_rubric_sheet(wb)
    build_axis_definitions_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"Generated {OUTPUT_PATH} with {len(TEST_PROMPTS)} test prompts")
    print(f"  - Single-slide: {sum(1 for p in TEST_PROMPTS if p[5] == '1')}")
    print(f"  - Multi-slide:  {sum(1 for p in TEST_PROMPTS if p[5] != '1')}")


if __name__ == "__main__":
    main()
