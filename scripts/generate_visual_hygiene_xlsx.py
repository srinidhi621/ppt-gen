"""
Generate the V3 visual hygiene checks Excel file.

These are objective, binary pass/fail checks for brand and mechanical
correctness. They test whether output follows template rules, NOT
whether it looks good, tells a story, or shows taste.

Every check is answerable by either:
  - the deterministic scanner (programmatic), or
  - a multimodal LLM looking at the rendered slide image (yes/no question), or
  - both.

Usage:
    python scripts/generate_visual_hygiene_xlsx.py

Output:
    assets/benchmarks/v3_visual_hygiene_checks.xlsx
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

OUTPUT_PATH = (
    Path(__file__).resolve().parent.parent
    / "assets" / "benchmarks" / "v3_visual_hygiene_checks.xlsx"
)

# ── Styles ───────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BLOCKING_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")


def style_header_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_body_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def add_section_row(ws, row, num_cols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BOLD_FONT
    cell.fill = SECTION_FILL
    cell.alignment = WRAP
    cell.border = THIN_BORDER
    for c in range(2, num_cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = SECTION_FILL


# ── Check definitions ────────────────────────────────────────────
# (check_id, category, check_name, llm_question, scanner_check,
#  pass_criteria, severity)

HYGIENE_CHECKS = [
    # ── Color ────────────────────────────────────────────────────
    (
        "VH-01", "Color", "All fills use palette tokens",
        "Are all colored shapes (rectangles, cards, bars, backgrounds) "
        "using colors from the brand palette? Look for any color that "
        "looks out of place or doesn't match the green/coral/navy/grey scheme.",
        "Yes — compare every shape fill RGB against token_overrides palette.",
        "Zero non-palette fills on non-image shapes.",
        "BLOCKING",
    ),
    (
        "VH-02", "Color", "All text colors use palette tokens",
        "Is all text rendered in brand-approved colors? Look for text "
        "that appears in an unexpected color (random blue, default black "
        "that should be charcoal, etc.).",
        "Yes — compare every text run font.color against token palette.",
        "Zero non-palette text colors.",
        "BLOCKING",
    ),
    (
        "VH-03", "Color", "Accent limit per slide",
        "Count the number of distinct accent colors on this slide "
        "(excluding neutral greys, white, and dark backgrounds). "
        "Are there more than 2?",
        "Yes — count distinct accent token roles used per slide.",
        "≤ 2 accent color roles per slide per accent_policy.",
        "WARNING",
    ),
    (
        "VH-04", "Color", "Text-background contrast",
        "Is any text hard to read because it's too close in color "
        "to its background? Look for light text on light backgrounds "
        "or dark text on dark backgrounds.",
        "Yes — compute contrast ratio per text run against underlying fill.",
        "Contrast ratio ≥ 4.5:1 for body text, ≥ 3:1 for large text (≥24pt).",
        "BLOCKING",
    ),
    (
        "VH-05", "Color", "No invisible shapes",
        "Are there any shapes whose fill color matches the slide "
        "background, making them invisible?",
        "Yes — compare shape fill against slide background/canvas fill.",
        "No shape has fill == background unless it is a deliberate text container.",
        "WARNING",
    ),
    # ── Typography ───────────────────────────────────────────────
    (
        "VH-06", "Typography", "All fonts from allowlist",
        "Are all text elements using the expected brand fonts "
        "(Space Grotesk / Inter or their originals PP Neue Machina / Aptos)? "
        "Look for any text in Calibri, Arial, Times New Roman, or other system defaults.",
        "Yes — compare every run font.name against font_substitution allowlist.",
        "Zero runs using non-allowlisted fonts.",
        "BLOCKING",
    ),
    (
        "VH-07", "Typography", "Font sizes match type scale",
        "Are text sizes drawn from a consistent scale? "
        "Look for text that's an awkward in-between size — "
        "not quite a heading, not quite body.",
        "Yes — compare every run font.size against type_scale defined steps "
        "(display 40pt, title 28pt, kicker 11pt, subtitle 16pt, body 12pt, caption 10pt).",
        "Every text run uses a type_scale step. Tolerance: ±1pt for measurement rounding.",
        "WARNING",
    ),
    (
        "VH-08", "Typography", "Bold matches type scale role",
        "Is bold used only where the type scale specifies it "
        "(titles, kickers, display)? Look for body text that's "
        "randomly bolded or titles that aren't bold.",
        "Yes — check bold attribute against type_scale role definition.",
        "Bold attribute matches the type_scale entry for that size/role.",
        "WARNING",
    ),
    (
        "VH-09", "Typography", "ALLCAPS only on kicker role",
        "Is there any text rendered in ALL CAPS that is not a kicker "
        "or section label? Look for body text or titles in all caps.",
        "Yes — check for uppercase text on runs not tagged as kicker role.",
        "Only kicker-role text may be uppercase (per type_scale.kicker.upper=true).",
        "WARNING",
    ),
    # ── Spatial ──────────────────────────────────────────────────
    (
        "VH-10", "Spatial", "No shapes outside canvas",
        "Are all shapes fully visible on the slide? "
        "Look for any text or shape that is cut off at the edges.",
        "Yes — check shape (left+width ≤ slide_width) and (top+height ≤ slide_height).",
        "Every shape bounding box fits within canvas dimensions.",
        "BLOCKING",
    ),
    (
        "VH-11", "Spatial", "No text overflow",
        "Is any text visibly cut off, truncated, or running outside "
        "its containing box? Look for sentences that end abruptly "
        "or text touching the edge of a shape.",
        "Yes — measure_text per text frame vs frame bounds.",
        "Measured text extent ≤ frame bounds for every text frame.",
        "BLOCKING",
    ),
    (
        "VH-12", "Spatial", "Content within safe area",
        "Is all content (text, shapes, images) within the slide margins? "
        "Look for elements that crowd the very edge of the slide.",
        "Yes — check shapes against safe_area from design_system.",
        "All content shapes within safe_area. Full-bleed backgrounds exempt.",
        "WARNING",
    ),
    (
        "VH-13", "Spatial", "No significant shape overlaps",
        "Do any content shapes (text boxes, cards, images) overlap each other "
        "in a way that makes content unreadable? Background shapes exempt.",
        "Yes — AABB intersection test with >10% overlap threshold.",
        "No two non-background shapes overlap by >10% of the smaller shape's area.",
        "WARNING",
    ),
    (
        "VH-14", "Spatial", "Consistent gutters",
        "Are the gaps between adjacent shapes (cards, columns, grid items) "
        "consistent? Look for uneven spacing between elements that should "
        "be equally spaced.",
        "Partial — can check if gaps match spacing_scale values.",
        "Gaps between peer shapes match a spacing_scale step (xs/sm/md/lg/xl). "
        "Tolerance: ±10% of the gap value.",
        "WARNING",
    ),
    (
        "VH-15", "Spatial", "Grid alignment",
        "Do shapes align to a visible column grid? "
        "Look for shapes that are slightly off-grid — shifted a few pixels "
        "left or right compared to their neighbors.",
        "Yes — check shape left positions against grid column boundaries.",
        "Shape left edges align to grid column starts within 0.05\" tolerance.",
        "WARNING",
    ),
    # ── Content Rendering ────────────────────────────────────────
    (
        "VH-16", "Content Rendering", "No leaked markdown",
        "Is there any visible markdown syntax on the slide? "
        "Look for literal **, *, ##, ```, or []() in the text.",
        "Yes — regex scan on all text runs for markdown patterns.",
        "Zero markdown syntax characters rendered as literal text.",
        "BLOCKING",
    ),
    (
        "VH-17", "Content Rendering", "No placeholder text",
        "Is there any template placeholder text visible? "
        "Look for {title}, {body}, Lorem ipsum, TODO, [insert], TBD, "
        "or similar placeholder patterns.",
        "Yes — regex scan for common placeholder patterns.",
        "Zero placeholder tokens in rendered text.",
        "BLOCKING",
    ),
    (
        "VH-18", "Content Rendering", "All images resolve",
        "Are all image placeholders filled with actual images? "
        "Look for broken image icons, empty picture frames, "
        "or red X markers.",
        "Yes — walk picture rels and verify package parts exist.",
        "Every picture relationship resolves to a valid image file.",
        "BLOCKING",
    ),
    (
        "VH-19", "Content Rendering", "No empty content frames",
        "Are there any visible text boxes or content areas that are "
        "completely empty (not decorative spacing, but content areas "
        "that should have text)?",
        "Partial — check text frames on non-background shapes for empty text.",
        "No content-role text frame has zero text runs.",
        "WARNING",
    ),
    # ── Cross-Slide Consistency ──────────────────────────────────
    (
        "VH-20", "Cross-Slide", "Title position consistent",
        "Looking at all slides in the deck, are titles in the same "
        "position on every slide? Look for titles that jump up/down "
        "or left/right between slides.",
        "Yes — compare title shape (left, top) across all slides.",
        "Title position varies by ≤ 0.1\" across all slides.",
        "WARNING",
    ),
    (
        "VH-21", "Cross-Slide", "Title style consistent",
        "Are all slide titles the same font, size, color, and weight? "
        "Look for a title that's suddenly smaller, different font, "
        "or different color.",
        "Yes — compare title font properties across all slides.",
        "Title font name, size, bold, and color are identical across all slides.",
        "BLOCKING",
    ),
    (
        "VH-22", "Cross-Slide", "Kicker style consistent",
        "If slides use kickers (small labels like '01 | The thesis'), "
        "are they all styled the same? Same font, size, position, casing?",
        "Yes — compare kicker-role text properties across slides that use them.",
        "Kicker font, size, color, casing, and vertical position are identical "
        "across all slides that use kickers.",
        "WARNING",
    ),
    (
        "VH-23", "Cross-Slide", "Body text style consistent",
        "Is body text the same font and size across all slides? "
        "Look for slides where the body text is noticeably larger "
        "or smaller than other slides.",
        "Yes — compare body-role text font properties across slides.",
        "Body text font name and size are identical across all slides. "
        "Color may vary by canvas (light vs dark).",
        "WARNING",
    ),
    # ── Structural ───────────────────────────────────────────────
    (
        "VH-24", "Structural", "No empty slides",
        "Are there any slides that appear completely blank or have "
        "only a background with no content?",
        "Yes — count non-placeholder shapes per slide.",
        "Every slide has ≥ 2 non-placeholder shapes.",
        "BLOCKING",
    ),
    (
        "VH-25", "Structural", "Slide count matches plan",
        "Does the number of slides match what the deck plan specified?",
        "Yes — compare len(slides) to deck_plan slide count.",
        "Slide count equals deck_plan.slides length.",
        "BLOCKING",
    ),
    (
        "VH-26", "Structural", "Visual element per content slide",
        "Does every content slide (not title/section break) have at "
        "least one non-text visual element — a shape, icon, image, "
        "colored rectangle, or line?",
        "Yes — check for at least one non-text-frame shape per content slide.",
        "Every content slide has ≥ 1 non-text visual element.",
        "WARNING",
    ),
]


# ── Excel generation ─────────────────────────────────────────────

def build_checks_sheet(wb):
    ws = wb.active
    ws.title = "Hygiene Checks"

    headers = [
        "Check ID", "Category", "Check Name",
        "LLM Visual Question (yes/no when viewing slide image)",
        "Scanner Automatable?",
        "Pass Criteria",
        "Severity",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    sections = [
        ("Color — are all colors from the brand palette?", "VH-01"),
        ("Typography — are fonts and sizes from the type scale?", "VH-06"),
        ("Spatial — do shapes fit the canvas and respect the grid?", "VH-10"),
        ("Content Rendering — is text rendered cleanly?", "VH-16"),
        ("Cross-Slide Consistency — does the deck look unified?", "VH-20"),
        ("Structural — is the deck well-formed?", "VH-24"),
    ]

    row = 2
    section_idx = 0
    for check in HYGIENE_CHECKS:
        cid = check[0]
        if section_idx < len(sections) and cid == sections[section_idx][1]:
            add_section_row(ws, row, len(headers), sections[section_idx][0])
            row += 1
            section_idx += 1

        for c, val in enumerate(check, 1):
            ws.cell(row=row, column=c, value=val)
        style_body_row(ws, row, len(headers))

        severity_cell = ws.cell(row=row, column=7)
        severity_cell.alignment = CENTER
        if check[6] == "BLOCKING":
            severity_cell.fill = BLOCKING_FILL
            severity_cell.font = Font(name="Calibri", size=10, bold=True, color="C62828")
        else:
            severity_cell.fill = WARNING_FILL
            severity_cell.font = Font(name="Calibri", size=10, bold=True, color="F57F17")

        row += 1

    # Summary
    row += 1
    blocking = sum(1 for c in HYGIENE_CHECKS if c[6] == "BLOCKING")
    warning = sum(1 for c in HYGIENE_CHECKS if c[6] == "WARNING")
    ws.cell(row=row, column=1, value="Summary:").font = BOLD_FONT
    ws.cell(row=row, column=2, value=(
        f"{len(HYGIENE_CHECKS)} checks total: {blocking} BLOCKING, {warning} WARNING. "
        "A deck passes visual hygiene if it has zero BLOCKING failures."
    )).font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["G"].width = 12

    for r in range(2, row):
        ws.row_dimensions[r].height = 65

    return ws


def build_scorecard_sheet(wb):
    ws = wb.create_sheet("Deck Scorecard")

    ws.cell(row=1, column=1, value="Deck Under Test:").font = BOLD_FONT
    ws.cell(row=1, column=2, value="(enter run_id or file path)").font = Font(
        name="Calibri", size=10, italic=True, color="999999"
    )
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

    headers = [
        "Check ID", "Check Name", "Severity",
        "Pass?", "Slide(s) Affected", "Notes",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    sections = [
        ("Color", "VH-01"),
        ("Typography", "VH-06"),
        ("Spatial", "VH-10"),
        ("Content Rendering", "VH-16"),
        ("Cross-Slide Consistency", "VH-20"),
        ("Structural", "VH-24"),
    ]

    row = 4
    section_idx = 0
    for check in HYGIENE_CHECKS:
        cid = check[0]
        if section_idx < len(sections) and cid == sections[section_idx][1]:
            add_section_row(ws, row, len(headers), sections[section_idx][0])
            row += 1
            section_idx += 1

        ws.cell(row=row, column=1, value=check[0])
        ws.cell(row=row, column=2, value=check[2])
        ws.cell(row=row, column=3, value=check[6])
        # Pass? column — user fills in YES/NO
        # Slide(s) Affected — user fills in
        # Notes — user fills in
        style_body_row(ws, row, len(headers))

        severity_cell = ws.cell(row=row, column=3)
        severity_cell.alignment = CENTER
        if check[6] == "BLOCKING":
            severity_cell.fill = BLOCKING_FILL
            severity_cell.font = Font(name="Calibri", size=10, bold=True, color="C62828")
        else:
            severity_cell.fill = WARNING_FILL
            severity_cell.font = Font(name="Calibri", size=10, bold=True, color="F57F17")

        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=5).alignment = WRAP
        ws.cell(row=row, column=6).alignment = WRAP

        row += 1

    # Summary formulas
    row += 1
    first_pass_row = 5  # first data row after header + first section
    last_pass_row = row - 2
    pass_col_letter = "D"

    ws.cell(row=row, column=1, value="Results:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Total Checks:").font = BODY_FONT
    ws.cell(row=row, column=3, value=len(HYGIENE_CHECKS)).font = BOLD_FONT
    ws.cell(row=row, column=3).alignment = CENTER

    row += 1
    ws.cell(row=row, column=2, value="Passed:").font = BODY_FONT
    ws.cell(row=row, column=3).font = BOLD_FONT
    ws.cell(row=row, column=3).alignment = CENTER
    ws.cell(row=row, column=3, value=(
        f'=COUNTIF({pass_col_letter}{first_pass_row}:{pass_col_letter}{last_pass_row},"YES")'
    ))

    row += 1
    ws.cell(row=row, column=2, value="Blocking Failures:").font = BODY_FONT
    ws.cell(row=row, column=3).font = Font(name="Calibri", size=10, bold=True, color="C62828")
    ws.cell(row=row, column=3).alignment = CENTER
    ws.cell(row=row, column=3, value=(
        f'=COUNTIFS({pass_col_letter}{first_pass_row}:{pass_col_letter}{last_pass_row},"NO",'
        f'C{first_pass_row}:C{last_pass_row},"BLOCKING")'
    ))

    row += 1
    ws.cell(row=row, column=2, value="Overall:").font = BODY_FONT
    ws.cell(row=row, column=3).font = BOLD_FONT
    ws.cell(row=row, column=3).alignment = CENTER
    blocking_failures_cell = f"C{row - 1}"
    ws.cell(row=row, column=3, value=(
        f'=IF({blocking_failures_cell}=0,"PASS","FAIL")'
    ))

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 40

    for r in range(4, row):
        if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < 28:
            ws.row_dimensions[r].height = 28

    return ws


def build_llm_prompt_sheet(wb):
    """Sheet with copy-pasteable LLM prompts for visual review."""
    ws = wb.create_sheet("LLM Review Prompts")

    ws.cell(row=1, column=1, value="Category").font = BOLD_FONT
    ws.cell(row=1, column=2, value="LLM Prompt (attach slide image and ask this)").font = BOLD_FONT
    style_header_row(ws, 1, 2)

    prompts = [
        (
            "Color (all checks)",
            "Look at this slide image. Answer each question YES or NO:\n"
            "1. Are all colored shapes using brand palette colors (greens, corals, navy, greys, white)? Any off-palette color?\n"
            "2. Is all text in brand-approved colors? Any text in an unexpected color?\n"
            "3. Are there more than 2 distinct accent colors on this slide?\n"
            "4. Is any text hard to read due to low contrast against its background?\n"
            "5. Are there any shapes that are invisible because they match the background?",
        ),
        (
            "Typography (all checks)",
            "Look at this slide image. Answer each question YES or NO:\n"
            "1. Is all text in a consistent brand font? Any text in Calibri, Arial, Times New Roman, or other system defaults?\n"
            "2. Do text sizes follow a clear scale (large titles, medium subtitles, smaller body)? Any text that's an awkward in-between size?\n"
            "3. Is bold used only for titles and labels, not randomly in body text?\n"
            "4. Is there any ALL CAPS text that is not a small label or kicker?",
        ),
        (
            "Spatial (all checks)",
            "Look at this slide image. Answer each question YES or NO:\n"
            "1. Are all shapes fully visible — nothing cut off at any edge?\n"
            "2. Is any text visibly cut off, truncated, or running outside its box?\n"
            "3. Is all content within reasonable margins (not crowding the slide edges)?\n"
            "4. Do any content elements overlap in a way that makes text unreadable?\n"
            "5. Are gaps between similar elements (cards, columns) consistent?\n"
            "6. Do shapes appear to align to a grid, or are some visibly off-alignment?",
        ),
        (
            "Content Rendering (all checks)",
            "Look at this slide image. Answer each question YES or NO:\n"
            "1. Is there any visible markdown syntax (**, *, ##, ```, []()) rendered as literal text?\n"
            "2. Is there any placeholder text like {title}, Lorem ipsum, TODO, [insert], or TBD?\n"
            "3. Are all image areas filled with actual images (no broken icons or empty frames)?\n"
            "4. Are there any content areas that appear completely empty?",
        ),
        (
            "Cross-Slide (all slides)",
            "Look at ALL slides in this deck. Answer each question YES or NO:\n"
            "1. Are slide titles in the same position on every slide?\n"
            "2. Are all titles the same font, size, and color?\n"
            "3. If kickers are used (small labels like '01 | Section'), are they styled consistently?\n"
            "4. Is body text the same font and size across all slides?",
        ),
        (
            "Structural (all slides)",
            "Look at ALL slides in this deck. Answer each question YES or NO:\n"
            "1. Are there any completely blank slides with no content?\n"
            "2. Does every content slide (not title/divider) have at least one visual element beyond text?",
        ),
    ]

    for r, (cat, prompt) in enumerate(prompts, 2):
        ws.cell(row=r, column=1, value=cat).font = BODY_FONT
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=prompt).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.row_dimensions[r].height = 120

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    return ws


def main():
    wb = openpyxl.Workbook()
    build_checks_sheet(wb)
    build_scorecard_sheet(wb)
    build_llm_prompt_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))

    blocking = sum(1 for c in HYGIENE_CHECKS if c[6] == "BLOCKING")
    warning = sum(1 for c in HYGIENE_CHECKS if c[6] == "WARNING")
    print(f"Generated {OUTPUT_PATH}")
    print(f"  {len(HYGIENE_CHECKS)} checks: {blocking} BLOCKING, {warning} WARNING")
    print(f"  3 sheets: Hygiene Checks, Deck Scorecard, LLM Review Prompts")


if __name__ == "__main__":
    main()
