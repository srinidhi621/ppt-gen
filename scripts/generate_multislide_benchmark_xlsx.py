"""
Generate the V3 multi-slide benchmark workbook.

This harness is separate from ``v3_test_prompts.xlsx``.  It is for deck-level
pipeline runs where the input should include enough per-slide source content to
test narrative planning, density handling, and cross-slide consistency without
giving the model an answer deck or direct source-deck link.

Usage:
    python scripts/generate_multislide_benchmark_xlsx.py

Output:
    assets/benchmarks/v3_multislide_test_prompts.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


OUTPUT_PATH = (
    Path(__file__).resolve().parent.parent
    / "assets"
    / "benchmarks"
    / "v3_multislide_test_prompts.xlsx"
)


MULTISLIDE_TESTS = [
    {
        "test_id": "MS-01",
        "category": "Legacy Modernization Pitch",
        "audience": "Bank CTO and Head of Engineering",
        "expected_slides": "6-7",
        "target_archetypes": (
            "hero_title, hero_statement_with_support_columns, comparison_split, "
            "process_flow, content_with_visual, timeline_roadmap"
        ),
        "deck_brief": (
            "Create a concise executive pitch deck for a legacy modernization "
            "program. The audience already knows modernization is needed, so the "
            "deck should make the case for why the work must be structured, safe, "
            "and measurable. Keep it executive-level, specific, and credible."
        ),
        "style_cues": (
            "Use a clean Ascendion-style business deck: high whitespace, sharp "
            "headlines, one strong idea per slide, restrained accent color, and "
            "simple diagrams instead of decorative illustrations. Avoid full black "
            "backgrounds and avoid dense paragraph cards."
        ),
        "slides": [
            {
                "title": "Legacy modernization needs a control system",
                "content": (
                    "Open with the program idea: Legacy Navigator. Explain that "
                    "the client needs to modernize core systems without losing "
                    "business behavior, security posture, or delivery velocity."
                ),
                "visual_cues": (
                    "Cover slide with a confident title treatment, small subtitle, "
                    "and a subtle system/navigation motif. Brand should remain visible."
                ),
            },
            {
                "title": "Fragmentation is slowing every release",
                "content": (
                    "Show the current-state problem: duplicate systems, undocumented "
                    "logic, manual reporting, brittle integrations, and 14-day change "
                    "lead time. Make the business cost clear without overdramatizing it."
                ),
                "visual_cues": (
                    "Before-state comparison or clustered friction map. Use 3-4 "
                    "labeled problem areas with one prominent operational metric."
                ),
            },
            {
                "title": "The target state is governed flow from intent to production",
                "content": (
                    "Describe the future state: unified platform boundaries, reusable "
                    "service contracts, real-time dashboards, automated validation, "
                    "and 3-day release cycles for priority changes."
                ),
                "visual_cues": (
                    "Two-sided current/future layout or left-to-right transformation "
                    "diagram. Keep the future side visibly simpler than the current side."
                ),
            },
            {
                "title": "Our approach reduces risk in staged gates",
                "content": (
                    "Use five stages: Discover, Assess, Stabilize, Migrate, Optimize. "
                    "Each stage should explain the work product, not just the activity. "
                    "Call out parity baseline, service boundaries, rollout safety, and "
                    "production learning loops."
                ),
                "visual_cues": (
                    "Horizontal process flow with numbered gates and short captions. "
                    "Use connectors and a bottom proof/risk-control line."
                ),
            },
            {
                "title": "Proof must connect delivery speed to operating quality",
                "content": (
                    "Use evidence points: 40% fewer incidents, deployment time cut "
                    "from two weeks to two days, 150 engineers onboarded, and $2.3M "
                    "annual savings. Do not add extra invented metrics."
                ),
                "visual_cues": (
                    "Open metric typography or compact KPI band. Numbers should be "
                    "dominant; labels should be short and scannable."
                ),
            },
            {
                "title": "The first 90 days should prove momentum",
                "content": (
                    "Lay out near-term next steps: schedule a workshop within two "
                    "weeks, run a technical deep-dive with the platform team, confirm "
                    "candidate domains, define the parity baseline, and submit proposal "
                    "by month end."
                ),
                "visual_cues": (
                    "Action roadmap or closing sequence with dates. Make ownership and "
                    "decision points visible without turning the slide into a checklist."
                ),
            },
        ],
        "evaluation_focus": (
            "Tests whether the system can build a 6-slide executive pitch from "
            "slide-level source content while preserving narrative flow and avoiding "
            "overlap in process/metric slides."
        ),
    },
    {
        "test_id": "MS-02",
        "category": "10x AI Engineering Operating Model",
        "audience": "Senior delivery, talent, and engineering leaders",
        "expected_slides": "7",
        "target_archetypes": (
            "hero_title, hero_statement_with_support_columns, process_flow, "
            "content_with_visual, comparison_split, timeline_roadmap"
        ),
        "deck_brief": (
            "Create a serious operating-model deck for a 10x AI engineering "
            "approach. The deck should explain the role profile, delivery system, "
            "pod packaging, outcomes, assessment funnel, level-by-level evaluation, "
            "and operational ownership. It should feel like a leadership working "
            "deck, not a marketing overview."
        ),
        "style_cues": (
            "Use a detailed but readable Ascendion-style operating-model design: "
            "structured stage cards, compact labels, color-coded gates, swimlane-like "
            "ownership cues, bottom summary bands, and one screenshot-style placeholder. "
            "Split dense content instead of shrinking text."
        ),
        "slides": [
            {
                "title": "10x AI engineers multiply delivery quality, not just speed",
                "content": (
                    "Open with the 10x AI Engineering Approach. Define the promise: "
                    "AI-augmented professionals who combine architecture ownership, "
                    "AI fluency, product judgment, and production discipline."
                ),
                "visual_cues": (
                    "Premium cover with bold 10x emphasis, a technical/AI visual motif, "
                    "and visible brand treatment. Avoid generic dark dashboard styling."
                ),
            },
            {
                "title": "The 10x profile blends six behaviors",
                "content": (
                    "Cover architecture-first mindset, AI-driven problem solving, "
                    "zoom-in/zoom-out judgment, systems thinking and collaboration, "
                    "fast learning and teaching, and security-first behavior."
                ),
                "visual_cues": (
                    "Six compact capability modules in a balanced grid or split across "
                    "two rows. Icons or small color accents are useful but not required."
                ),
            },
            {
                "title": "Delivery runs from architecture intent to safe production",
                "content": (
                    "Show three connected stages: Architecture to Code, Code Generation, "
                    "and Code to Production. Include spec contract, legacy parity baseline, "
                    "approved generation, validation gates, rollout safety, and feedback "
                    "from production learnings."
                ),
                "visual_cues": (
                    "Three-stage connected process with clear arrows or handoff cues. "
                    "Each stage should have 3-4 short proof points."
                ),
            },
            {
                "title": "10x pods package premium capability around outcomes",
                "content": (
                    "Explain pod composition, pricing, engagement model, and scope. "
                    "Composition: 1-2 AI architecture or practice leads plus 3-5 "
                    "AI-augmented engineers. Pricing: premium offshore band of 10-15K "
                    "per resource per month. Scope includes SDD-style development, "
                    "AI-assisted code generation, testing, CI/CD, and observability."
                ),
                "visual_cues": (
                    "Four-part operating model. Use a visible pricing emphasis but keep "
                    "the slide from becoming a price card."
                ),
            },
            {
                "title": "The expected outcomes must be measurable",
                "content": (
                    "Cover faster time-to-value, higher quality and reliability, AI "
                    "maturity uplift, ROI, and economic advantage. Tie each outcome to "
                    "delivery mechanics such as shorter feedback loops, test coverage, "
                    "security checks, auditability, and reduced rework."
                ),
                "visual_cues": (
                    "Stacked outcome callouts beside a simple business-impact visual or "
                    "signal chart. Keep the outcome labels short."
                ),
            },
            {
                "title": "The campus funnel tests skill, judgment, and fit",
                "content": (
                    "Show six assessment stages: HackerRank test, business problem, "
                    "solution presentation, HR or org fit, psychometric test, and offer "
                    "rollout. Include durations: 120 minutes, 24-hour take-home, 45-minute "
                    "panel, 30-minute HR, 30-60 minute psychometric, final offer."
                ),
                "visual_cues": (
                    "Horizontal funnel or staged roadmap with durations and short criteria. "
                    "Add a lower lane for Talent Teams, Technical Panels, HR, Eightfold, "
                    "and HackerRank."
                ),
            },
            {
                "title": "Each level needs a clear signal and owner",
                "content": (
                    "Summarize Level 1 HackerRank, Level 2 business problem/system design, "
                    "Level 3 solution presentation or technical panel, final HR assessment, "
                    "offer rollout, and operational mechanics. Include platforms Eightfold "
                    "and HackerRank; owners Talent Acquisition, business leaders, technical "
                    "panel, and HR. End with the operating principle: one funnel, four "
                    "owners, two platforms."
                ),
                "visual_cues": (
                    "Dense but controlled ownership matrix or split process summary. Use a "
                    "bottom band for 'Operating principle' and avoid tiny paragraphs."
                ),
            },
        ],
        "evaluation_focus": (
            "Tests whether the system can convert detailed slide-level operating-model "
            "content into a 7-slide deck without needing access to the reference deck."
        ),
    },
]


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _headers() -> list[str]:
    headers = [
        "Test ID",
        "Category",
        "Audience",
        "Expected Slides",
        "Target Archetype(s)",
        "Deck Brief",
        "Style / Visual System Cues",
    ]
    for slide_num in range(1, 8):
        headers.extend([
            f"Slide {slide_num} Content",
            f"Slide {slide_num} Visual Cues",
        ])
    headers.append("Evaluation Focus")
    return headers


def _row_for_test(test: dict) -> list[str]:
    row = [
        test["test_id"],
        test["category"],
        test["audience"],
        test["expected_slides"],
        test["target_archetypes"],
        test["deck_brief"],
        test["style_cues"],
    ]
    slides = test["slides"]
    for index in range(7):
        if index < len(slides):
            slide = slides[index]
            row.extend([
                f"{slide['title']}\n{slide['content']}",
                slide["visual_cues"],
            ])
        else:
            row.extend(["", ""])
    row.append(test["evaluation_focus"])
    return row


def _style_row(ws, row_num: int, num_cols: int, *, header: bool = False) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT if header else BODY_FONT
        cell.alignment = CENTER if header else WRAP
        cell.border = THIN_BORDER
        if header:
            cell.fill = HEADER_FILL


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Multi-Slide Prompts"

    headers = _headers()
    ws.append(headers)
    _style_row(ws, 1, len(headers), header=True)

    for test in MULTISLIDE_TESTS:
        ws.append(_row_for_test(test))
        _style_row(ws, ws.max_row, len(headers))
        ws.row_dimensions[ws.max_row].height = 150

    widths = {
        "A": 10,
        "B": 26,
        "C": 28,
        "D": 14,
        "E": 32,
        "F": 44,
        "G": 44,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for col in range(8, 22):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 36
    ws.column_dimensions["V"].width = 44

    note = wb.create_sheet("Harness Notes")
    note.append(["Purpose", "Guidance"])
    _style_row(note, 1, 2, header=True)
    note.append([
        "Why this exists",
        (
            "This workbook is for multi-slide pipeline tests. It gives source "
            "content and ambiguous visual cues per slide, but does not link to or "
            "describe an answer deck."
        ),
    ])
    note.append([
        "How to run",
        (
            "Use scripts/run_v3_eval_prompts.py --harness multislide. The runner "
            "composes each row into a markdown-style user instruction for V3."
        ),
    ])
    for row_num in range(2, note.max_row + 1):
        _style_row(note, row_num, 2)
        note.row_dimensions[row_num].height = 80
    note.column_dimensions["A"].width = 24
    note.column_dimensions["B"].width = 100
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    if OUTPUT_PATH.exists():
        OUTPUT_PATH.unlink()
    wb.save(str(OUTPUT_PATH))
    print(f"Generated {OUTPUT_PATH} with {len(MULTISLIDE_TESTS)} multi-slide tests")


if __name__ == "__main__":
    main()
